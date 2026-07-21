"""
NovaChem Knowledge Graph - Graph Builder (Team B / Person 2)
=================================================================
Builds a NetworkX graph from multiple sources:
  1. equipment_master.xlsx        -> Equipment nodes (with attributes)
  2. equipment_relationships.xlsx -> Equipment-to-equipment edges (process flow)
  3. entities_final.json          -> Everything Qwen extracted from documents
                                      (FailureMode, Person, Standard nodes + edges)
  4. plant_events.xlsx            -> Authoritative event log (grounded=True)
  5. plant_layout.xlsx            -> Spatial 'near' relationships
  6. novachem_employees.xlsx      -> Person enrichment
  7. failure_taxonomy.xlsx        -> Classification hierarchy
  8. documents.jsonl              -> Public Reference nodes

CHANGE LOG (duplicate-edge refactor):
--------------------------------------
* Added add_unique_edge() helper — replaces ALL direct G.add_edge() calls.
  This is the ONLY place edges are now inserted.
  If an edge (src, relation, tgt) already exists:
      - source_docs list is extended (no duplicates)
      - event_ids list is extended (no duplicates)
      - grounded is promoted to True if any source says True
      - occurrence_count is incremented
  Otherwise a new edge is created with occurrence_count=1.

* Every loader function that previously called G.add_edge() now calls
  add_unique_edge() instead. Zero other logic is changed.

Usage:
    python build_graph.py \\
        --equipment_master Dataset/equipment_master.xlsx \\
        --equipment_relationships Dataset/equipment_relationships.xlsx \\
        --entities entities_final.json \\
        --out graph.gpickle

Requires: networkx, openpyxl
    pip install networkx openpyxl
"""

import argparse
import json
import pickle

import networkx as nx
import openpyxl


# ---------------------------------------------------------------------------
# NEW: Central edge-insertion helper
# ---------------------------------------------------------------------------

def add_unique_edge(
    G: nx.MultiDiGraph,
    source: str,
    target: str,
    relation: str,
    source_doc: str | None = None,
    event_id: str | None = None,
    grounded: bool = True,
    **extra_attrs,
) -> None:
    """
    Insert an edge (source, relation, target) into G exactly once.

    If an edge with the same (source, target, relation) already exists,
    its metadata is merged rather than creating a duplicate parallel edge:

      - source_docs  : list — new source_doc appended if not already present
      - event_ids    : list — new event_id appended if not already present
      - grounded     : bool — promoted to True if ANY occurrence is grounded
      - occurrence_count : int — incremented on every call

    If no matching edge exists, a fresh edge is created with
    occurrence_count=1.

    Parameters
    ----------
    G          : the MultiDiGraph being built
    source     : source node id (must already exist in G)
    target     : target node id (must already exist in G)
    relation   : relation label string (e.g. "has_failure")
    source_doc : originating file path / dataset name (optional)
    event_id   : event identifier if edge comes from an event (optional)
    grounded   : whether this occurrence is verified in the source text
    **extra_attrs : any extra edge attributes (e.g. reason, doc_type) —
                    only written on edge creation, not on merge, because
                    they are static facts that don't change per occurrence.
    """
    # --- Search existing parallel edges between (source, target) ----------
    # G[source][target] is a dict keyed by integer edge-keys for MultiDiGraph.
    # We scan all existing parallel edges to find one with matching relation.
    existing_key = None
    if G.has_node(source) and G.has_node(target) and target in G[source]:
        for edge_key, edge_data in G[source][target].items():
            if edge_data.get("relation") == relation:
                existing_key = edge_key
                break

    if existing_key is not None:
        # --- MERGE into the existing edge -----------------------------------
        edge_data = G[source][target][existing_key]

        # Append new source_doc to the list if it is genuinely new
        if source_doc and source_doc not in edge_data["source_docs"]:
            edge_data["source_docs"].append(source_doc)

        # Append new event_id to the list if it is genuinely new
        if event_id and event_id not in edge_data["event_ids"]:
            edge_data["event_ids"].append(event_id)

        # Once grounded by any source, always grounded
        if grounded:
            edge_data["grounded"] = True

        # Track how many documents contributed this same fact
        edge_data["occurrence_count"] += 1

    else:
        # --- CREATE a new edge ---------------------------------------------
        G.add_edge(
            source,
            target,
            relation=relation,
            # Store provenance as lists from the start so the merge path
            # never has to deal with scalar → list conversion.
            source_docs=[source_doc] if source_doc else [],
            event_ids=[event_id] if event_id else [],
            grounded=grounded,
            occurrence_count=1,
            **extra_attrs,
        )


# ---------------------------------------------------------------------------
# Loader functions — unchanged logic, only G.add_edge → add_unique_edge
# ---------------------------------------------------------------------------

def load_equipment_master(path: str, G: nx.MultiDiGraph):
    """Load equipment nodes from equipment_master.xlsx.
    No edges are created here — unchanged from original."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    for row in rows[1:]:
        record = dict(zip(headers, row))
        eq_id = record.get("Equipment ID")
        if not eq_id:
            continue
        G.add_node(eq_id, node_type="Equipment",
                   **{k: v for k, v in record.items() if v is not None})


def load_plant_events(path: str, G: nx.MultiDiGraph):
    """
    Load authoritative event log from plant_events.xlsx.

    CHANGE: All five G.add_edge() calls replaced with add_unique_edge().
    This means the same failure mode appearing in multiple events will
    result in ONE edge with occurrence_count > 1 instead of N parallel edges.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        eq_id    = r.get("Equipment ID")
        event_id = r.get("Event ID")
        if not eq_id or not event_id:
            continue

        if eq_id not in G:
            G.add_node(eq_id, node_type="Equipment")

        chain_node = f"EVENT:{event_id}"
        # EventChain node: created or attributes updated in place
        G.add_node(chain_node, node_type="EventChain",
                   date=str(r.get("Date")), severity=r.get("Severity"),
                   priority=r.get("Priority"), status=r.get("Status"))

        # CHANGED: was G.add_edge(chain_node, eq_id, relation="involves", ...)
        add_unique_edge(
            G, chain_node, eq_id,
            relation="involves",
            source_doc="plant_events.xlsx",
            grounded=True,
        )

        description       = r.get("Description")
        root_cause        = r.get("Root Cause")
        corrective_action = r.get("Corrective Action")
        reported_by       = r.get("Reported By")
        assigned_to       = r.get("Assigned To")

        if description:
            G.add_node(description, node_type="FailureMode")
            # CHANGED: was G.add_edge(eq_id, description, relation="has_failure", ...)
            add_unique_edge(
                G, eq_id, description,
                relation="has_failure",
                source_doc="plant_events.xlsx",
                event_id=event_id,
                grounded=True,
            )
            if root_cause:
                G.add_node(root_cause, node_type="FailureMode")
                # CHANGED: was G.add_edge(description, root_cause, relation="caused_by", ...)
                add_unique_edge(
                    G, description, root_cause,
                    relation="caused_by",
                    source_doc="plant_events.xlsx",
                    event_id=event_id,
                    grounded=True,
                )

        if corrective_action:
            G.add_node(corrective_action, node_type="MaintenanceAction")
            # CHANGED: was G.add_edge(eq_id, corrective_action, relation="resolved_by", ...)
            add_unique_edge(
                G, eq_id, corrective_action,
                relation="resolved_by",
                source_doc="plant_events.xlsx",
                event_id=event_id,
                grounded=True,
            )

        for person, rel in ((reported_by, "reported_by"), (assigned_to, "assigned_to")):
            if person:
                if person not in G:
                    G.add_node(person, node_type="Person")
                # CHANGED: was G.add_edge(eq_id, person, relation=rel, ...)
                add_unique_edge(
                    G, eq_id, person,
                    relation=rel,
                    source_doc="plant_events.xlsx",
                    event_id=event_id,
                    grounded=True,
                )


def load_failure_taxonomy(path: str, G: nx.MultiDiGraph):
    """
    Load failure taxonomy hierarchy from failure_taxonomy.xlsx.

    CHANGE: All three G.add_edge() calls replaced with add_unique_edge().
    Taxonomy edges are structural (not per-document) so occurrence_count
    will always stay at 1 here, but the merge logic is harmless.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        l1, l2, l3 = r.get("Level 1"), r.get("Level 2"), r.get("Level 3")
        if not l3:
            continue
        G.add_node(l3, node_type="FailureTaxonomy", level1=l1, level2=l2,
                   possible_cause=r.get("Possible Cause"),
                   affected_equipment=r.get("Affected Equipment"),
                   recommended_action=r.get("Recommended Action"))
        if l2:
            G.add_node(l2, node_type="FailureTaxonomy", level1=l1)
            # CHANGED: was G.add_edge(l3, l2, relation="part_of", ...)
            add_unique_edge(
                G, l3, l2,
                relation="part_of",
                source_doc="failure_taxonomy.xlsx",
                grounded=True,
            )
        if l1:
            G.add_node(l1, node_type="FailureTaxonomy")
            if l2:
                # CHANGED: was G.add_edge(l2, l1, relation="part_of", ...)
                add_unique_edge(
                    G, l2, l1,
                    relation="part_of",
                    source_doc="failure_taxonomy.xlsx",
                    grounded=True,
                )

        # Link to a real extracted FailureMode node if name matches closely.
        # Substring match — unchanged from original.
        for node, data in list(G.nodes(data=True)):
            if data.get("node_type") == "FailureMode" and l3.lower() in str(node).lower():
                # CHANGED: was G.add_edge(node, l3, relation="classified_as", ...)
                add_unique_edge(
                    G, node, l3,
                    relation="classified_as",
                    source_doc="failure_taxonomy.xlsx",
                    grounded=True,
                )


def load_employees(path: str, G: nx.MultiDiGraph):
    """
    Enrich existing Person nodes (or create them) from novachem_employees.xlsx.
    No edges are created here — unchanged from original.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        name = r.get("Name")
        if not name:
            continue
        attrs = {k: v for k, v in r.items() if v is not None and k != "Name"}
        if name in G:
            G.nodes[name].update(attrs)
            G.nodes[name]["node_type"] = "Person"
        else:
            G.add_node(name, node_type="Person", **attrs)


def load_plant_layout(path: str, G: nx.MultiDiGraph):
    """
    Add spatial 'near' edges from plant_layout.xlsx.

    CHANGE: G.add_edge() replaced with add_unique_edge().
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        eq_id = r.get("Equipment ID")
        if not eq_id:
            continue
        attrs = {
            "area": r.get("Area"), "building": r.get("Building"),
            "floor": r.get("Floor"), "criticality": r.get("Criticality"),
        }
        if eq_id in G:
            G.nodes[eq_id].update({k: v for k, v in attrs.items() if v is not None})
        else:
            G.add_node(eq_id, node_type="Equipment", **attrs)

        nearby = r.get("Nearby Equipment")
        if nearby:
            for n in str(nearby).split(","):
                n = n.strip()
                if n and n in G:
                    # CHANGED: was G.add_edge(eq_id, n, relation="near", ...)
                    add_unique_edge(
                        G, eq_id, n,
                        relation="near",
                        source_doc="plant_layout.xlsx",
                        grounded=True,
                    )


def load_public_references(documents_jsonl_path: str, G: nx.MultiDiGraph):
    """
    Add public Reference nodes from documents.jsonl (SDS/manuals/regulations).

    CHANGE: G.add_edge() replaced with add_unique_edge().
    """
    aliases = {}
    manufacturers = {}
    for node, data in G.nodes(data=True):
        if data.get("node_type") == "Equipment":
            if data.get("Equipment Name"):
                aliases[str(data["Equipment Name"]).strip().lower()] = node
            if data.get("Manufacturer"):
                manufacturers.setdefault(
                    str(data["Manufacturer"]).lower(), []
                ).append(node)

    linked_count = 0
    with open(documents_jsonl_path, encoding="utf-8") as f:
        for line in f:
            doc = json.loads(line)
            if doc.get("layer") != "public_industrial_knowledge":
                continue

            ref_node = doc["file_path"]
            G.add_node(ref_node, node_type="Reference",
                       doc_type=doc.get("doc_type"),
                       file_name=doc.get("file_name"))

            text_lower = doc.get("text", "")[:3000].lower()

            matched_equipment = set()
            for eq_name, eq_id in aliases.items():
                if eq_name in text_lower:
                    matched_equipment.add(eq_id)
            for mfr_name, eq_ids in manufacturers.items():
                if mfr_name in text_lower:
                    matched_equipment.update(eq_ids)

            for eq_id in matched_equipment:
                # CHANGED: was G.add_edge(eq_id, ref_node, relation="has_reference", ...)
                add_unique_edge(
                    G, eq_id, ref_node,
                    relation="has_reference",
                    source_doc="keyword_match",
                    grounded=True,
                )
                linked_count += 1

    print(
        f"Added {sum(1 for _, d in G.nodes(data=True) if d.get('node_type') == 'Reference')} "
        f"Reference nodes, {linked_count} links to equipment"
    )


def load_equipment_relationships(path: str, G: nx.MultiDiGraph):
    """
    Load process-flow edges from equipment_relationships.xlsx.

    CHANGE: G.add_edge() replaced with add_unique_edge().
    The 'reason' field is passed as an extra_attr (written on creation only).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    for row in rows[1:]:
        record = dict(zip(headers, row))
        src = record.get("Source Equipment")
        rel = record.get("Relationship")
        tgt = record.get("Target Equipment")
        if src and rel and tgt:
            # CHANGED: was G.add_edge(src, tgt, relation=rel, reason=..., source=...)
            add_unique_edge(
                G, src, tgt,
                relation=rel,
                source_doc="equipment_relationships.xlsx",
                grounded=True,
                reason=record.get("Reason"),   # extra_attr, written on creation
            )


def load_extracted_entities(path: str, G: nx.MultiDiGraph):
    """
    Load entities and relationships extracted by Qwen from entities_final.json.

    CHANGE: G.add_edge() replaced with add_unique_edge().
    This is the highest-volume call site — it is responsible for most of
    the duplicate edges in the original implementation, because the same
    (equipment, failure_mode) pair appears in every document for that event.
    """
    with open(path, encoding="utf-8") as f:
        documents = json.load(f)

    for doc in documents:
        doc_ref  = doc.get("file_path")
        event_id = doc.get("event_id")

        # Add non-equipment entities as nodes (unchanged)
        for e in doc.get("entities", []):
            name, etype = e.get("name"), e.get("type")
            if not name:
                continue
            if name not in G:
                G.add_node(name, node_type=etype)
            # Track every document this node is mentioned in (unchanged)
            mentions = G.nodes[name].setdefault("mentioned_in", set())
            if doc_ref:
                mentions.add(doc_ref)

        for r in doc.get("relationships", []):
            src, rel, tgt = r.get("source"), r.get("relation"), r.get("target")
            if not (src and rel and tgt):
                continue
            if src not in G:
                G.add_node(src, node_type="Unknown")
            if tgt not in G:
                G.add_node(tgt, node_type="Unknown")
            # CHANGED: was G.add_edge(src, tgt, relation=rel, source=doc_ref, ...)
            add_unique_edge(
                G, src, tgt,
                relation=rel,
                source_doc=doc_ref,
                event_id=event_id,
                grounded=r.get("grounded", True),
            )


def link_same_event_documents(G: nx.MultiDiGraph, entities_path: str):
    """
    Add lightweight EventChain nodes and connect all entities per event.

    CHANGE: G.add_edge() replaced with add_unique_edge().
    The involves edges here have no source_doc because link_same_event_documents
    doesn't track individual file paths — consistent with original behaviour.
    """
    with open(entities_path, encoding="utf-8") as f:
        documents = json.load(f)

    event_to_entities: dict[str, set] = {}
    for doc in documents:
        event_id = doc.get("event_id")
        if not event_id:
            continue
        event_to_entities.setdefault(event_id, set())
        for e in doc.get("entities", []):
            if e.get("name"):
                event_to_entities[event_id].add(e["name"])

    for event_id, entity_names in event_to_entities.items():
        chain_node = f"EVENT:{event_id}"
        G.add_node(chain_node, node_type="EventChain")
        for name in entity_names:
            if name in G:
                # CHANGED: was G.add_edge(chain_node, name, relation="involves")
                add_unique_edge(
                    G, chain_node, name,
                    relation="involves",
                    # No source_doc — matches original behaviour for these edges
                )


# ---------------------------------------------------------------------------
# Internal query helpers (used only during build-time demo, unchanged)
# ---------------------------------------------------------------------------

def get_failure_history(G: nx.MultiDiGraph, equipment_id: str) -> list:
    """All failure modes ever linked to a piece of equipment."""
    if equipment_id not in G:
        return []
    failures = []
    for _, target, data in G.out_edges(equipment_id, data=True):
        if data.get("relation") == "has_failure":
            failures.append({
                "failure_mode":    target,
                # CHANGED: field is now source_docs (list), not source (str)
                "source_docs":     data.get("source_docs", []),
                "event_ids":       data.get("event_ids", []),
                "grounded":        data.get("grounded", True),
                "occurrence_count": data.get("occurrence_count", 1),
            })
    return failures


def get_related_incidents(G: nx.MultiDiGraph, equipment_id: str) -> list:
    """Every event/document this equipment was mentioned in. Unchanged."""
    if equipment_id not in G:
        return []
    events = []
    for pred in G.predecessors(equipment_id):
        if G.nodes[pred].get("node_type") == "EventChain":
            events.append(pred.replace("EVENT:", ""))
    return events


def get_standards_for(G: nx.MultiDiGraph, equipment_id: str) -> list:
    """Unchanged."""
    if equipment_id not in G:
        return []
    return [t for _, t, d in G.out_edges(equipment_id, data=True)
            if d.get("relation") == "follows_standard"]


def get_who_worked_on(G: nx.MultiDiGraph, equipment_id: str) -> list:
    """Unchanged."""
    if equipment_id not in G:
        return []
    return [t for _, t, d in G.out_edges(equipment_id, data=True)
            if d.get("relation") in ("repaired_by", "documented_by")]


def get_upstream_downstream(G: nx.MultiDiGraph, equipment_id: str) -> dict:
    """Unchanged — uses source_doc field name which is now source_docs (list).
    Updated to check membership in the list."""
    if equipment_id not in G:
        return {"upstream": [], "downstream": []}
    upstream = [
        u for u, _, d in G.in_edges(equipment_id, data=True)
        if "equipment_relationships.xlsx" in d.get("source_docs", [])
    ]
    downstream = [
        v for _, v, d in G.out_edges(equipment_id, data=True)
        if "equipment_relationships.xlsx" in d.get("source_docs", [])
    ]
    return {"upstream": upstream, "downstream": downstream}


# ---------------------------------------------------------------------------
# main()
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--equipment_master",       required=True)
    parser.add_argument("--equipment_relationships", required=True)
    parser.add_argument("--entities",               required=True)
    parser.add_argument("--plant_events",     default=None,
                        help="Path to plant_events.xlsx (authoritative event log)")
    parser.add_argument("--failure_taxonomy", default=None,
                        help="Path to failure_taxonomy.xlsx")
    parser.add_argument("--employees",        default=None,
                        help="Path to novachem_employees.xlsx")
    parser.add_argument("--plant_layout",     default=None,
                        help="Path to plant_layout.xlsx")
    parser.add_argument("--documents_jsonl",  default=None,
                        help="Path to documents.jsonl — adds public Reference nodes")
    parser.add_argument("--out", default="data/graph/knowledge_graph.gpickle")
    args = parser.parse_args()

    G = nx.MultiDiGraph()
    load_equipment_master(args.equipment_master, G)
    load_equipment_relationships(args.equipment_relationships, G)

    if args.plant_events:
        load_plant_events(args.plant_events, G)
        print("Loaded plant_events.xlsx (authoritative event log)")
    if args.plant_layout:
        load_plant_layout(args.plant_layout, G)
        print("Loaded plant_layout.xlsx (spatial relationships)")
    if args.employees:
        load_employees(args.employees, G)
        print("Loaded novachem_employees.xlsx (person enrichment)")

    load_extracted_entities(args.entities, G)
    link_same_event_documents(G, args.entities)

    if args.failure_taxonomy:
        load_failure_taxonomy(args.failure_taxonomy, G)
        print("Loaded failure_taxonomy.xlsx (classification hierarchy)")
    if args.documents_jsonl:
        load_public_references(args.documents_jsonl, G)

    print(f"Graph built: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")

    from collections import Counter
    type_counts = Counter(nx.get_node_attributes(G, "node_type").values())
    print("\nNode types:")
    for t, c in type_counts.most_common():
        print(f"  {t}: {c}")

    # Convert mentioned_in sets to lists before pickling (unchanged)
    for n, data in G.nodes(data=True):
        if isinstance(data.get("mentioned_in"), set):
            data["mentioned_in"] = list(data["mentioned_in"])

    with open(args.out, "wb") as f:
        pickle.dump(G, f)
    print(f"\nSaved graph to {args.out}")

    # Quick demo query on a known equipment ID
    demo_id = "P-101"
    if demo_id in G:
        print(f"\n--- Demo query: {demo_id} ---")
        print("Failure history:", get_failure_history(G, demo_id)[:3])
        print("Standards:",       get_standards_for(G, demo_id))
        print("Upstream/downstream:", get_upstream_downstream(G, demo_id))


if __name__ == "__main__":
    main()
