"""
NovaChem Knowledge Graph - Graph Builder (Team B / Person 2)
=================================================================
Builds a NetworkX graph from three sources:
  1. equipment_master.xlsx        -> Equipment nodes (with attributes)
  2. equipment_relationships.xlsx -> Equipment-to-equipment edges (process flow)
  3. entities_final.json          -> Everything Qwen extracted from documents
                                      (FailureMode, Person, Standard nodes + edges)

Usage:
    python build_graph.py \\
        --equipment_master Dataset/equipment_master.xlsx \\
        --equipment_relationships Dataset/equipment_relationships.xlsx \\
        --entities entities_final.json \\
        --out graph.gpickle

Requires: networkx, openpyxl
    pip install networkx openpyxl
"""

import argparse
import json
import pickle

import networkx as nx
import openpyxl


def load_equipment_master(path: str, G: nx.DiGraph):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    for row in rows[1:]:
        record = dict(zip(headers, row))
        eq_id = record.get("Equipment ID")
        if not eq_id:
            continue
        G.add_node(eq_id, node_type="Equipment", **{k: v for k, v in record.items() if v is not None})


def load_plant_events(path: str, G: nx.DiGraph):
    """plant_events.xlsx is the authoritative master log - every fact here is
    100% trusted (no LLM involved), so all edges get grounded=True and
    source='plant_events.xlsx'. This directly fills in has_failure, caused_by,
    corrective actions, and who reported/was assigned - the same relationship
    types Qwen was extracting from prose, except these are ground truth."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        eq_id = r.get("Equipment ID")
        event_id = r.get("Event ID")
        if not eq_id or not event_id:
            continue

        if eq_id not in G:
            G.add_node(eq_id, node_type="Equipment")

        chain_node = f"EVENT:{event_id}"
        G.add_node(chain_node, node_type="EventChain",
                    date=str(r.get("Date")), severity=r.get("Severity"),
                    priority=r.get("Priority"), status=r.get("Status"))
        G.add_edge(chain_node, eq_id, relation="involves", source="plant_events.xlsx")

        description = r.get("Description")
        root_cause = r.get("Root Cause")
        corrective_action = r.get("Corrective Action")
        reported_by = r.get("Reported By")
        assigned_to = r.get("Assigned To")

        if description:
            G.add_node(description, node_type="FailureMode")
            G.add_edge(eq_id, description, relation="has_failure", source="plant_events.xlsx",
                       event_id=event_id, grounded=True)
            if root_cause:
                G.add_node(root_cause, node_type="FailureMode")
                G.add_edge(description, root_cause, relation="caused_by", source="plant_events.xlsx",
                           event_id=event_id, grounded=True)

        if corrective_action:
            G.add_node(corrective_action, node_type="MaintenanceAction")
            G.add_edge(eq_id, corrective_action, relation="resolved_by", source="plant_events.xlsx",
                       event_id=event_id, grounded=True)

        for person, rel in ((reported_by, "reported_by"), (assigned_to, "assigned_to")):
            if person:
                if person not in G:
                    G.add_node(person, node_type="Person")
                G.add_edge(eq_id, person, relation=rel, source="plant_events.xlsx",
                           event_id=event_id, grounded=True)


def load_failure_taxonomy(path: str, G: nx.DiGraph):
    """Reference hierarchy (Mechanical -> Bearing Failure -> Bearing Wear) with
    recommended actions - links to existing FailureMode nodes by name match
    where possible, otherwise stands alone as browsable reference data."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        l1, l2, l3 = r.get("Level 1"), r.get("Level 2"), r.get("Level 3")
        if not l3:
            continue
        G.add_node(l3, node_type="FailureTaxonomy", level1=l1, level2=l2,
                   possible_cause=r.get("Possible Cause"),
                   affected_equipment=r.get("Affected Equipment"),
                   recommended_action=r.get("Recommended Action"))
        if l2:
            G.add_node(l2, node_type="FailureTaxonomy", level1=l1)
            G.add_edge(l3, l2, relation="part_of", source="failure_taxonomy.xlsx", grounded=True)
        if l1:
            G.add_node(l1, node_type="FailureTaxonomy")
            if l2:
                G.add_edge(l2, l1, relation="part_of", source="failure_taxonomy.xlsx", grounded=True)

        # link to a real extracted FailureMode node if the name matches closely
        for node, data in list(G.nodes(data=True)):
            if data.get("node_type") == "FailureMode" and l3.lower() in str(node).lower():
                G.add_edge(node, l3, relation="classified_as", source="failure_taxonomy.xlsx", grounded=True)


def load_employees(path: str, G: nx.DiGraph):
    """Enriches existing Person nodes (or creates them) with real department/
    role/experience data instead of just a bare name."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        name = r.get("Name")
        if not name:
            continue
        attrs = {k: v for k, v in r.items() if v is not None and k != "Name"}
        if name in G:
            G.nodes[name].update(attrs)
            G.nodes[name]["node_type"] = "Person"
        else:
            G.add_node(name, node_type="Person", **attrs)


def load_plant_layout(path: str, G: nx.DiGraph):
    """Adds physical location attributes and spatial 'near' edges between
    equipment - relationship data you don't get from process-flow or incidents."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]

    for row in rows[1:]:
        r = dict(zip(headers, row))
        eq_id = r.get("Equipment ID")
        if not eq_id:
            continue
        attrs = {"area": r.get("Area"), "building": r.get("Building"), "floor": r.get("Floor"),
                 "criticality": r.get("Criticality")}
        if eq_id in G:
            G.nodes[eq_id].update({k: v for k, v in attrs.items() if v is not None})
        else:
            G.add_node(eq_id, node_type="Equipment", **attrs)

        nearby = r.get("Nearby Equipment")
        if nearby:
            for n in str(nearby).split(","):
                n = n.strip()
                if n and n in G:
                    G.add_edge(eq_id, n, relation="near", source="plant_layout.xlsx", grounded=True)


def load_public_references(documents_jsonl_path: str, G: nx.DiGraph):
    """Public documents (SDS/manuals/regulations/brochures) don't fit the
    Equipment/FailureMode/Person/Standard ontology - full entity extraction on
    them produced 138 invented entity types. Instead, add each one as a single
    lightweight Reference node and link it to real equipment only where there's
    a genuine, checkable match (equipment name/manufacturer mentioned in the
    document), rather than trying to extract structured facts from them."""
    aliases = {}
    manufacturers = {}
    for node, data in G.nodes(data=True):
        if data.get("node_type") == "Equipment":
            if data.get("Equipment Name"):
                aliases[str(data["Equipment Name"]).strip().lower()] = node
            if data.get("Manufacturer"):
                manufacturers.setdefault(str(data["Manufacturer"]).lower(), []).append(node)

    linked_count = 0
    with open(documents_jsonl_path, encoding="utf-8") as f:
        for line in f:
            doc = json.loads(line)
            if doc.get("layer") != "public_industrial_knowledge":
                continue

            ref_node = doc["file_path"]
            G.add_node(ref_node, node_type="Reference", doc_type=doc.get("doc_type"),
                       file_name=doc.get("file_name"))

            text_lower = doc.get("text", "")[:3000].lower()  # first ~3000 chars is enough for this check

            matched_equipment = set()
            for eq_name, eq_id in aliases.items():
                if eq_name in text_lower:
                    matched_equipment.add(eq_id)
            for mfr_name, eq_ids in manufacturers.items():
                if mfr_name in text_lower:
                    matched_equipment.update(eq_ids)

            for eq_id in matched_equipment:
                G.add_edge(eq_id, ref_node, relation="has_reference", source="keyword_match", grounded=True)
                linked_count += 1

    print(f"Added {sum(1 for _, d in G.nodes(data=True) if d.get('node_type') == 'Reference')} "
          f"Reference nodes, {linked_count} links to equipment")


def load_equipment_relationships(path: str, G: nx.DiGraph):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    for row in rows[1:]:
        record = dict(zip(headers, row))
        src, rel, tgt = record.get("Source Equipment"), record.get("Relationship"), record.get("Target Equipment")
        if src and rel and tgt:
            G.add_edge(src, tgt, relation=rel, reason=record.get("Reason"), source="equipment_relationships.xlsx")


def load_extracted_entities(path: str, G: nx.DiGraph):
    with open(path, encoding="utf-8") as f:
        documents = json.load(f)

    for doc in documents:
        doc_ref = doc.get("file_path")
        event_id = doc.get("event_id")

        # add non-equipment entities as nodes (equipment nodes should already exist
        # from equipment_master; if not, add them so nothing gets dropped silently)
        for e in doc.get("entities", []):
            name, etype = e.get("name"), e.get("type")
            if not name:
                continue
            if name not in G:
                G.add_node(name, node_type=etype)
            # track every document/event a node is mentioned in
            mentions = G.nodes[name].setdefault("mentioned_in", set())
            if doc_ref:
                mentions.add(doc_ref)

        for r in doc.get("relationships", []):
            src, rel, tgt = r.get("source"), r.get("relation"), r.get("target")
            if not (src and rel and tgt):
                continue
            if src not in G:
                G.add_node(src, node_type="Unknown")
            if tgt not in G:
                G.add_node(tgt, node_type="Unknown")
            G.add_edge(src, tgt, relation=rel, source=doc_ref, event_id=event_id,
                       grounded=r.get("grounded", True))


def link_same_event_documents(G: nx.DiGraph, entities_path: str):
    """Adds a lightweight EventChain node per event_id, connecting every entity
    that was mentioned in a document belonging to that event."""
    with open(entities_path, encoding="utf-8") as f:
        documents = json.load(f)

    event_to_entities = {}
    for doc in documents:
        event_id = doc.get("event_id")
        if not event_id:
            continue
        event_to_entities.setdefault(event_id, set())
        for e in doc.get("entities", []):
            if e.get("name"):
                event_to_entities[event_id].add(e["name"])

    for event_id, entity_names in event_to_entities.items():
        chain_node = f"EVENT:{event_id}"
        G.add_node(chain_node, node_type="EventChain")
        for name in entity_names:
            if name in G:
                G.add_edge(chain_node, name, relation="involves")


# ---------------------------------------------------------------------------
# Query functions - what the RAG Copilot will actually call
# ---------------------------------------------------------------------------

def get_failure_history(G: nx.DiGraph, equipment_id: str) -> list:
    """All failure modes ever linked to a piece of equipment."""
    if equipment_id not in G:
        return []
    failures = []
    for _, target, data in G.out_edges(equipment_id, data=True):
        if data.get("relation") == "has_failure":
            failures.append({
                "failure_mode": target,
                "source_doc": data.get("source"),
                "event_id": data.get("event_id"),
                "grounded": data.get("grounded", True),
            })
    return failures


def get_related_incidents(G: nx.DiGraph, equipment_id: str) -> list:
    """Every event/document this equipment was mentioned in."""
    if equipment_id not in G:
        return []
    events = []
    for pred in G.predecessors(equipment_id):
        if G.nodes[pred].get("node_type") == "EventChain":
            events.append(pred.replace("EVENT:", ""))
    return events


def get_standards_for(G: nx.DiGraph, equipment_id: str) -> list:
    if equipment_id not in G:
        return []
    return [t for _, t, d in G.out_edges(equipment_id, data=True) if d.get("relation") == "follows_standard"]


def get_who_worked_on(G: nx.DiGraph, equipment_id: str) -> list:
    if equipment_id not in G:
        return []
    return [t for _, t, d in G.out_edges(equipment_id, data=True)
            if d.get("relation") in ("repaired_by", "documented_by")]


def get_upstream_downstream(G: nx.DiGraph, equipment_id: str) -> dict:
    """Process-flow neighbors from equipment_relationships.xlsx (e.g. what
    feeds into this equipment, what it feeds into)."""
    if equipment_id not in G:
        return {"upstream": [], "downstream": []}
    upstream = [u for u, _, d in G.in_edges(equipment_id, data=True) if d.get("source") == "equipment_relationships.xlsx"]
    downstream = [v for _, v, d in G.out_edges(equipment_id, data=True) if d.get("source") == "equipment_relationships.xlsx"]
    return {"upstream": upstream, "downstream": downstream}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--equipment_master", required=True)
    parser.add_argument("--equipment_relationships", required=True)
    parser.add_argument("--entities", required=True)
    parser.add_argument("--plant_events", default=None, help="Path to plant_events.xlsx (authoritative event log)")
    parser.add_argument("--failure_taxonomy", default=None, help="Path to failure_taxonomy.xlsx")
    parser.add_argument("--employees", default=None, help="Path to novachem_employees.xlsx")
    parser.add_argument("--plant_layout", default=None, help="Path to plant_layout.xlsx")
    parser.add_argument("--documents_jsonl", default=None,
                         help="Path to documents.jsonl (from read_documents.py) - if given, adds public "
                              "documents (SDS/manuals/regulations) as lightweight Reference nodes")
    parser.add_argument("--out", default="graph.gpickle")
    args = parser.parse_args()

    G = nx.MultiDiGraph()
    load_equipment_master(args.equipment_master, G)
    load_equipment_relationships(args.equipment_relationships, G)

    if args.plant_events:
        load_plant_events(args.plant_events, G)
        print("Loaded plant_events.xlsx (authoritative event log)")
    if args.plant_layout:
        load_plant_layout(args.plant_layout, G)
        print("Loaded plant_layout.xlsx (spatial relationships)")
    if args.employees:
        load_employees(args.employees, G)
        print("Loaded novachem_employees.xlsx (person enrichment)")

    load_extracted_entities(args.entities, G)
    link_same_event_documents(G, args.entities)

    if args.failure_taxonomy:
        load_failure_taxonomy(args.failure_taxonomy, G)
        print("Loaded failure_taxonomy.xlsx (classification hierarchy)")

    if args.documents_jsonl:
        load_public_references(args.documents_jsonl, G)

    print(f"Graph built: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")

    from collections import Counter
    type_counts = Counter(nx.get_node_attributes(G, "node_type").values())
    print("\nNode types:")
    for t, c in type_counts.most_common():
        print(f"  {t}: {c}")

    # sets aren't picklable cleanly with some protocols - convert mentioned_in back to lists
    for n, data in G.nodes(data=True):
        if isinstance(data.get("mentioned_in"), set):
            data["mentioned_in"] = list(data["mentioned_in"])

    with open(args.out, "wb") as f:
        pickle.dump(G, f)
    print(f"\nSaved graph to {args.out}")

    # quick demo query on a known equipment ID
    demo_id = "P-101"
    if demo_id in G:
        print(f"\n--- Demo query: {demo_id} ---")
        print("Failure history:", get_failure_history(G, demo_id)[:3])
        print("Standards:", get_standards_for(G, demo_id))
        print("Upstream/downstream:", get_upstream_downstream(G, demo_id))


if __name__ == "__main__":
    main()
